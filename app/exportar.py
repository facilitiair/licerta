"""Exportação dos resultados filtrados em CSV e XLSX (SPEC §7)."""
import csv
import io
import re

# O Excel proíbe caracteres de controle — e o PNCP às vezes os envia no objeto
_ILEGAIS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

COLUNAS = [
    ("numero_controle_pncp", "Nº controle PNCP"),
    ("modalidade_nome", "Modalidade"),
    ("orgao_nome", "Órgão"),
    ("unidade_nome", "Unidade"),
    ("municipio_nome", "Município"),
    ("uf", "UF"),
    ("objeto", "Objeto"),
    ("valor_total_estimado", "Valor estimado (R$)"),
    ("srp", "SRP"),
    ("numero_compra", "Nº compra"),
    ("ano_compra", "Ano"),
    ("processo", "Processo"),
    ("data_publicacao_pncp", "Publicação"),
    ("data_abertura_proposta", "Abertura propostas"),
    ("data_encerramento_proposta", "Encerramento propostas"),
    ("link_pncp", "Link PNCP"),
    ("link_sistema_origem", "Sistema de origem"),
]


def _linha(lic):
    valores = []
    for campo, _ in COLUNAS:
        v = getattr(lic, campo, None)
        if campo == "srp":
            v = "sim" if v else "não"
        if isinstance(v, str):
            v = _ILEGAIS.sub(" ", v)
        valores.append("" if v is None else v)
    return valores


def gerar_csv(licitacoes):
    """CSV com ; e BOM — abre direto no Excel brasileiro."""
    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    escritor.writerow([titulo for _, titulo in COLUNAS])
    for lic in licitacoes:
        escritor.writerow(_linha(lic))
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def gerar_xlsx(licitacoes):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    aba = wb.active
    aba.title = "Licitações"
    aba.append([titulo for _, titulo in COLUNAS])
    for lic in licitacoes:
        aba.append(_linha(lic))
    for i, (campo, _) in enumerate(COLUNAS, 1):
        aba.column_dimensions[get_column_letter(i)].width = \
            60 if campo == "objeto" else 22
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
