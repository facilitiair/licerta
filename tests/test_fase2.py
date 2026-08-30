"""Testes da Fase 2: exportação CSV/XLSX e HTML do alerta por e-mail."""
import io

from app.alerta import _texto_para_html
from app.exportar import COLUNAS, gerar_csv, gerar_xlsx


class LicFake:
    numero_controle_pncp = "x-1-000001/2026"
    modalidade_nome = "Pregão - Eletrônico"
    orgao_nome = "PREFEITURA DE TESTE"
    unidade_nome = "SECRETARIA"
    municipio_nome = "Teresina"
    uf = "PI"
    objeto = "Objeto com acentuação; e ponto-e-vírgula"
    valor_total_estimado = 1234.56
    srp = True
    numero_compra = "1"
    ano_compra = 2026
    processo = "123/2026"
    data_publicacao_pncp = "2026-08-30T10:00:00"
    data_abertura_proposta = "2026-09-01T08:00:00"
    data_encerramento_proposta = "2026-09-10T09:00:00"
    link_pncp = "https://pncp.gov.br/app/editais/x/2026/1"
    link_sistema_origem = "https://origem.exemplo"


def test_csv_tem_cabecalho_bom_e_escapa_ponto_e_virgula():
    dados = gerar_csv([LicFake()])
    texto = dados.decode("utf-8-sig")          # BOM presente para o Excel
    linhas = texto.splitlines()
    assert linhas[0].startswith("Nº controle PNCP;Modalidade")
    assert '"Objeto com acentuação; e ponto-e-vírgula"' in linhas[1]
    assert "sim" in linhas[1]                  # srp formatado


def test_xlsx_valido_com_todas_as_colunas():
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(gerar_xlsx([LicFake(), LicFake()])))
    aba = wb.active
    assert aba.max_row == 3                    # cabeçalho + 2 linhas
    assert aba.max_column == len(COLUNAS)
    assert aba.cell(2, 7).value.startswith("Objeto com")


def test_html_do_email_escapa_e_quebra_linhas():
    html = _texto_para_html("linha 1\nlinha <2>")
    assert "linha 1<br>" in html
    assert "&lt;2&gt;" in html                 # HTML da licitação não vaza tags
